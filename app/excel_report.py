"""Generation du rapport Excel de demo (factures, releve, rapprochement)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.demo_data import DemoBankLine, DemoInvoice
from app.reconciliation import ReconciliationResult
from app.vat import simulate_vat


def build_excel_report(
    invoices: list[DemoInvoice],
    bank_lines: list[DemoBankLine],
    reconciliations: list[ReconciliationResult],
    output_path: str,
) -> str:
    wb = Workbook()

    ws_inv = wb.active
    ws_inv.title = "Factures (demo)"
    ws_inv.append(["Fournisseur", "Numero", "Date", "HT", "Taux TVA %", "TVA", "TTC", "Categorie"])
    for cell in ws_inv[1]:
        cell.font = Font(bold=True)
    for inv in invoices:
        vat = simulate_vat(inv.montant_ht, inv.taux_tva)
        ws_inv.append(
            [
                inv.fournisseur,
                inv.numero,
                inv.date_facture.isoformat(),
                float(vat.montant_ht),
                float(vat.taux_tva),
                float(vat.montant_tva),
                float(vat.montant_ttc),
                "demo",
            ]
        )

    ws_bank = wb.create_sheet("Releve bancaire (demo)")
    ws_bank.append(["Date", "Libelle", "Montant"])
    for cell in ws_bank[1]:
        cell.font = Font(bold=True)
    for line in bank_lines:
        ws_bank.append([line.date_operation.isoformat(), line.libelle, float(line.montant)])

    ws_rec = wb.create_sheet("Rapprochement (demo)")
    ws_rec.append(["Facture", "Statut", "Detail"])
    for cell in ws_rec[1]:
        cell.font = Font(bold=True)
    for rec in reconciliations:
        ws_rec.append([rec.invoice.numero, rec.status, rec.detail])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
