from pathlib import Path

from openpyxl import load_workbook

from app.demo_data import generate_demo_bank_statement, generate_demo_invoices
from app.excel_report import build_excel_report
from app.reconciliation import reconcile_invoices


def test_build_excel_report_creates_three_sheets(tmp_path: Path):
    invoices = generate_demo_invoices(count=3, seed=42)
    bank_lines = generate_demo_bank_statement(invoices, seed=1)
    reconciliations = reconcile_invoices(invoices, bank_lines)

    output_path = tmp_path / "rapport.xlsx"
    build_excel_report(invoices, bank_lines, reconciliations, str(output_path))

    assert output_path.exists()
    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == {"Factures (demo)", "Releve bancaire (demo)", "Rapprochement (demo)"}
    ws_inv = wb["Factures (demo)"]
    assert ws_inv.max_row == len(invoices) + 1  # + en-tete
